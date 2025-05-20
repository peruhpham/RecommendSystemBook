from tkinter import *
from tkinter import messagebox, filedialog
from tkinter import ttk
import pandas as pd
import openpyxl
import csv

# Node chứa thông tin của một quyển sách
class Book:
    def __init__(self, title=None, author=None, year=None):
        self.title = title
        self.author = author
        self.year = year
        self.next = None  # Con trỏ đến sách tiếp theo

# Danh sách liên kết đơn để quản lý sách
class BookList:
    def __init__(self):
        self.head = None  # Điểm bắt đầu của danh sách

    def append(self, title, author, year):
        new_book = Book(title, author, year)
        if not self.head:
            self.head = new_book
        else:
            current = self.head
            while current.next:
                current = current.next
            current.next = new_book

    def to_list(self):
        books = []
        current = self.head
        while current:
            books.append({"Title": current.title, "Author": current.author, "Year": current.year})
            current = current.next
        return books

    def display(self, listbox):
        listbox.delete(0, END)
        current = self.head
        while current:
            listbox.insert(END, f"{current.title} - {current.author} ({current.year})")
            current = current.next

class BookDataManagement:
    def __init__(self, parent):
        self.parent = parent
        self.books = BookList()  # Danh sách liên kết để lưu trữ sách
        self.create_ui_data()

        # Đường dẫn đến tệp CSV
        # file_path = "D:\\dataBookRecommend\\Books.csv"  # Thay bằng đường dẫn tệp của bạn
        file_path = "D:\Recommend-system-book-streamlit-Sentence-Transformers\books-test.csv"  # Thay bằng đường dẫn tệp của bạn

        # Tải dữ liệu từ CSV vào danh sách liên kết
        load_books_from_csv(file_path, self.books)

        # Hiển thị dữ liệu từ danh sách liên kết vào Treeview
        load_treeview_from_booklist(self.treeview, self.books)

    def create_ui_data(self):
        self.ui_data_window = Toplevel(self.parent)
        self.ui_data_window.title("UI Data Management")
        self.ui_data_window.geometry("920x400")
        # Đặt icon cho cửa sổ
        try:
            self.icon_image = PhotoImage(file="images/icon.png")  # Đảm bảo đường dẫn đúng
            self.ui_data_window.iconphoto(False, self.icon_image)  # Gọi iconphoto trên Toplevel
        except Exception as e:
            print(f"Lỗi đặt icon: {e}")
        self.ui_data_window.config(bg="#F0F0F0")

        self.style = ttk.Style(self.ui_data_window)
        try:
            self.ui_data_window.tk.call("source", r"setup\tkinter_excel_app\forest-dark.tcl")
            self.ui_data_window.tk.call("source", r"setup\tkinter_excel_app\forest-light.tcl")
        except TclError as e:
            print(f"Lỗi tcl: {e}")

        self.style.theme_use("forest-dark")

        # Tạo nút chuyển đổi chế độ sáng/tối
        # self.mode_switch = ttk.Checkbutton(
        #     self.ui_data_window,
        #     text="Chế độ sáng/tối",
        #     command=self.toggle_mode,
        # )
        # self.mode_switch.state(["!alternate"])  # Đảm bảo trạng thái mặc định
        # self.mode_switch.pack(anchor="ne", padx=0, pady=0)

        # Frame chính
        frame = ttk.Frame(self.ui_data_window)
        frame.pack(fill=BOTH, expand=True, padx=10, pady=10)

        widgets_frame = ttk.LabelFrame(frame, text="Quản lý dữ liệu")
        widgets_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        self.name_entry = ttk.Entry(widgets_frame)
        self.name_entry.insert(0, "Name book")
        self.name_entry.bind("<FocusIn>", lambda e: self.name_entry.delete(0, 'end'))
        self.name_entry.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

        self.author_spinbox = ttk.Spinbox(widgets_frame, from_=18, to=100)
        self.author_spinbox.insert(0, "Author")
        self.author_spinbox.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

        self.combo_list = ["*****", "****", "***", "**", "*"]
        self.status_combobox = ttk.Combobox(widgets_frame, values=self.combo_list)
        self.status_combobox.current(0)
        self.status_combobox.grid(row=2, column=0, padx=5, pady=5, sticky="ew")

        self.employed_var = BooleanVar()
        self.checkbutton = ttk.Checkbutton(widgets_frame, text="Employed", variable=self.employed_var)
        self.checkbutton.grid(row=3, column=0, padx=5, pady=5, sticky="nsew")

        insert_button = ttk.Button(widgets_frame, text="Thêm", command=self.insert_to_treeview)
        insert_button.grid(row=4, column=0, padx=5, pady=5, sticky="nsew")

        delete_button = ttk.Button(widgets_frame, text="Xóa", command=self.delete_selected_row)
        delete_button.grid(row=5, column=0, padx=5, pady=5, sticky="nsew")

        export_button = ttk.Button(widgets_frame, text="Xuất Excel", command=self.export_to_excel)
        export_button.grid(row=6, column=0, padx=5, pady=5, sticky="nsew")

        tree_frame = ttk.Frame(frame)
        tree_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

        tree_scroll = ttk.Scrollbar(tree_frame)
        tree_scroll.pack(side="right", fill="y")

        # cols = ("Name", "Author", "Rating", "Genre")
        # self.treeview = ttk.Treeview(tree_frame, show="headings", columns=cols, yscrollcommand=tree_scroll.set)
        # for col in cols:
        #     self.treeview.heading(col, text=col)
        #     self.treeview.column(col, width=150)
        cols = ("STT", "Name", "Author", "Year", "Rating")
        self.treeview = ttk.Treeview(tree_frame, show="headings", columns=cols, yscrollcommand=tree_scroll.set)
        # Cấu hình tiêu đề và kích thước cột
        self.treeview.heading("STT", text="STT")
        self.treeview.column("STT", width=50, anchor="center")  # Cột số thứ tự nhỏ hơn và căn giữa
        for col in cols[1:]:  # Cấu hình các cột còn lại
            self.treeview.heading(col, text=col)
            self.treeview.column(col, width=150)
        self.treeview.pack(fill=BOTH, expand=True)
        tree_scroll.config(command=self.treeview.yview)

    def toggle_mode(self):
        if self.mode_switch.instate(["selected"]):
            self.style.theme_use("forest-light")
        else:
            self.style.theme_use("forest-dark")

    def insert_to_treeview(self):
        name = self.name_entry.get().strip()
        author = self.author_spinbox.get().strip()
        subscription = self.status_combobox.get().strip()
        employment_status = "Yes" if self.employed_var.get() else "No"

        if not name or not author:
            messagebox.showwarning("Cảnh báo", "Vui lòng nhập tên và tuổi hợp lệ!")
            return

        self.treeview.insert("", "end", values=(name, author, subscription, employment_status))
        self.name_entry.delete(0, "end")
        self.name_entry.insert(0, "Name")
        self.author_spinbox.delete(0, "end")
        self.author_spinbox.insert(0, "Author")
        self.status_combobox.current(0)
        self.checkbutton.state(["!selected"])

    def delete_selected_row(self):
        selected_item = self.treeview.selection()
        if not selected_item:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn hàng cần xóa!")
            return
        self.treeview.delete(selected_item)

    def export_to_excel(self):
        """
        Xuất dữ liệu từ Treeview ra tệp Excel (.xlsx).
        """
        # Chọn nơi lưu file
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not path:
            return  # Nếu không chọn file, thoát hàm

        # Kiểm tra dữ liệu trong Treeview
        if not self.treeview.get_children():
            messagebox.showwarning("Cảnh báo", "Không có dữ liệu để xuất!")
            return

        try:
            # Tạo workbook và sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Data"

            # Thêm tiêu đề cột
            cols = ("Name", "Author", "Year","Rating", "Genres")
            for col_num, col_name in enumerate(cols, start=1):
                cell = sheet.cell(row=1, column=col_num, value=col_name)
                cell.font = openpyxl.styles.Font(bold=True)  # In đậm tiêu đề
                cell.alignment = openpyxl.styles.Alignment(horizontal="center")  # Căn giữa

            # Thêm dữ liệu từ Treeview
            for row_num, child in enumerate(self.treeview.get_children(), start=2):
                values = self.treeview.item(child)["values"]
                for col_num, value in enumerate(values, start=1):
                    sheet.cell(row=row_num, column=col_num, value=value)

            # Lưu file
            workbook.save(path)
            messagebox.showinfo("Thông báo", f"Dữ liệu đã được xuất ra tệp Excel tại {path}")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xuất dữ liệu: {e}")

def load_books_from_csv(file_path, book_list):
    """
    Đọc dữ liệu từ tệp CSV và thêm vào danh sách liên kết.
    """
    try:
        with open(file_path, "r", encoding="utf-8") as file:
            reader = csv.reader(file)
            next(reader)  # Bỏ qua dòng tiêu đề
            for row in reader:
                title = row[1]  # Cột "Book-Title"
                author = row[2]  # Cột "Book-Author"
                year = row[3]  # Cột "Year-Of-Publication"
                book_list.append(title, author, year)
        print("✅ Dữ liệu book.csv đã được tải vào danh sách liên kết!")
    except Exception as e:
        print(f"❌ Lỗi khi đọc tệp CSV: {e}")

def load_treeview_from_booklist(treeview, book_list):
    """
    Hiển thị dữ liệu từ danh sách liên kết vào Treeview.
    """
    treeview.delete(*treeview.get_children())  # Xóa tất cả dữ liệu cũ trong Treeview
    current = book_list.head
    while current:
        # print(f"Thêm vào Treeview: {current.title}, {current.author}, {current.year}")
        treeview.insert("", "end", values=(current.title, current.author, current.year))
        current = current.next

# def load_data_to_treeview(treeview, file_path):
#     """
#     Đọc dữ liệu từ tệp CSV và hiển thị vào Treeview.
#     """
#     try:
#         with open(file_path, "r", encoding="utf-8") as file:
#             reader = csv.reader(file)
#             next(reader)  # Bỏ qua dòng tiêu đề
#             treeview.delete(*treeview.get_children())  # Xóa dữ liệu cũ trong Treeview
#             for index, row in enumerate(reader, start=1):
#                 title = row[1]  # Cột "Book-Title"
#                 author = row[2]  # Cột "Book-Author"
#                 year = row[3]  # Cột "Year-Of-Publication"
#                 treeview.insert("", "end", values=(index, title, author, year))
#         print("✅ Dữ liệu đã được tải vào Treeview!")
#     except Exception as e:
#         print(f"❌ Lỗi khi đọc tệp CSV: {e}")

# def load_data_to_treeview(treeview, books_file, ratings_file):
#     """
#     Đọc dữ liệu từ Books.csv và Ratings.csv, sau đó hiển thị vào Treeview.
#     """
#     try:
#         # Tải dữ liệu Ratings.csv
#         ratings = load_ratings(ratings_file)

#         # Đọc dữ liệu từ Books.csv
#         with open(books_file, "r", encoding="utf-8") as file:
#             reader = csv.reader(file)
#             next(reader)  # Bỏ qua dòng tiêu đề
#             treeview.delete(*treeview.get_children())  # Xóa dữ liệu cũ trong Treeview
#             for index, row in enumerate(reader, start=1):
#                 isbn = row[0].strip()  # Cột "ISBN"
#                 title = row[1].strip()
#                 author = row[2].strip()
#                 year = row[3].strip()
#                 rating = ratings.get(isbn, "N/A")  # Lấy Rating từ ánh xạ, nếu không có thì "N/A"
#                 treeview.insert("", "end", values=(index, title, author, year, rating))
#         print("✅ Dữ liệu đã được tải vào Treeview!")
#     except Exception as e:
#         print(f"❌ Lỗi khi đọc tệp Books.csv: {e}")

def load_data_to_treeview(treeview, books_file, ratings_file):
    try:
        # Tải dữ liệu Ratings.csv
        ratings = load_ratings(ratings_file)

        # Đọc dữ liệu từ Books.csv
        with open(books_file, "r", encoding="utf-8") as file:
            reader = csv.reader(file)
            next(reader)  # Bỏ qua dòng tiêu đề
            treeview.delete(*treeview.get_children())  # Xóa dữ liệu cũ trong Treeview
            for index, row in enumerate(reader, start=1):
                isbn = row[0].strip()  # Loại bỏ khoảng trắng
                title = row[1]
                author = row[2]
                year = row[3]
                rating = ratings.get(isbn, "N/A")  # Lấy Rating từ ánh xạ, nếu không có thì "N/A"
                print(f"ISBN={isbn}, Rating={rating}")  # Kiểm tra ánh xạ
                treeview.insert("", "end", values=(index, title, author, year, rating))
        print("✅ Dữ liệu đã được tải vào Treeview!")
    except Exception as e:
        print(f"❌ Lỗi khi đọc tệp Books.csv: {e}")


def load_data_to_treeview(treeview, books_file):
    try:
        # Đọc dữ liệu từ Books.csv
        with open(books_file, "r", encoding="utf-8") as file:
            reader = csv.reader(file)
            next(reader)  # Bỏ qua dòng tiêu đề
            treeview.delete(*treeview.get_children())  # Xóa dữ liệu cũ trong Treeview
            for index, row in enumerate(reader, start=1):
                title = row[0]
                author = row[1]
                year = row[2]
                rating = row[3]
                description = row[4]
                image_url = row[5]
                # Kiểm tra dữ liệu trước khi thêm vào Treeview
                if not title or not author or not year:
                    print(f"❌ Dữ liệu không hợp lệ: {row}")
                    continue
                # Thêm vào Treeview
                treeview.insert("", "end", values=(index, title, author, year, rating, description, image_url))
        print("✅ Dữ liệu đã được tải vào Treeview!")
    except Exception as e:
        print(f"❌ Lỗi khi đọc tệp Books.csv: {e}")


def load_ratings(file_path):
    ratings = {}
    try:
        with open(file_path, "r", encoding="utf-8") as file:
            reader = csv.reader(file)
            next(reader)  # Bỏ qua dòng tiêu đề
            for row in reader:
                isbn = row[0].strip()  # Loại bỏ khoảng trắng
                rating = row[1].strip()
                ratings[isbn] = rating
                print(f"Đã thêm: ISBN={isbn}, Rating={rating}")
        print("✅ Dữ liệu Ratings.csv đã được tải!")
    except Exception as e:
        print(f"❌ Lỗi khi đọc tệp Ratings.csv: {e}")
    return ratings
