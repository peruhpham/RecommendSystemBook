from tkinter import *
from tkinter import ttk  # Import ttk cho các widget nâng cao
from tkinter import messagebox
from models.utils import create_rounded_rectangle
from controllers.book_management import BookDataManagement
from controllers.book_management import load_data_to_treeview
import requests
from PIL import Image, ImageTk
from io import BytesIO
import urllib.parse
import time
from models.utils import create_book_display
import openpyxl
from tkinter import filedialog


class BookRecommendationSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Book Recommendation System")
        # self.root.config(bg="#555555")
        self.root.config(bg="white")

        # Căn giữa cửa sổ
        self.center_window(1250, 670)

        self.inc = 0
        self.check_var_date = BooleanVar()
        self.check_var_rating = BooleanVar()
        self.employed_var = BooleanVar()  # Sửa: Lưu biến làm thuộc tính của lớp

        self.create_ui()

        # Frame bên trái
        self.create_left_frame()

        # # Tải dữ liệu từ tệp CSV vào Treeview
        # file_path = "D:\\dataBookRecommend\\Books.csv"  # Thay bằng đường dẫn tệp của bạn
        # load_data_to_treeview(self.treeview, file_path)


        # # Tải dữ liệu từ tệp CSV vào Treeview
        # books_file = "D:\\dataBookRecommend\\Books.csv"  # Thay bằng đường dẫn tệp của bạn
        # ratings_file = "D:\\dataBookRecommend\\Ratings.csv"  # Thay bằng đường dẫn tệp của bạn
        # load_data_to_treeview(self.treeview, books_file, ratings_file)
        
        # Tải dữ liệu từ tệp CSV vào Treeview
        books_file = r"D:\Recommend-system-book-streamlit-Sentence-Transformers\books-test.csv"

        # ratings_file = "D:\\dataBookRecommend\\Ratings.csv"  # Thay bằng đường dẫn tệp của bạn
        load_data_to_treeview(self.treeview, books_file)  # Chỉ cần truyền tệp sách, không cần tệp đánh giá

    def create_left_frame(self):
        """Tạo frame bên trái với các thành phần quản lý dữ liệu."""
        left_frame = ttk.LabelFrame(self.root, text="Quản lý dữ liệu")
        left_frame.place(x=20, y=410, width=200, height=250)  # Đặt tại tọa độ (20, 300)

        # Các thành phần trong frame bên trái
        name_label = ttk.Label(left_frame, text="Tên sách:")
        name_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.name_entry = ttk.Entry(left_frame)
        self.name_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        author_label = ttk.Label(left_frame, text="Tác giả:")
        author_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.author_spinbox = ttk.Spinbox(left_frame, from_=1, to=100, width=15)
        self.author_spinbox.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        rating_label = ttk.Label(left_frame, text="Đánh giá:")
        rating_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")
        combo_list = ["*****", "****", "***", "**", "*"]
        self.rating_combobox = ttk.Combobox(left_frame, values=combo_list)
        self.rating_combobox.current(0)
        self.rating_combobox.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

        # Sửa: Sử dụng biến thuộc tính của lớp
        self.checkbutton = ttk.Checkbutton(left_frame, text="Employed", variable=self.employed_var)
        self.checkbutton.grid(row=3, column=0, columnspan=2, padx=5, pady=5, sticky="w")

        # Các nút chức năng
        self.insert_button = ttk.Button(left_frame, text="Thêm", command=self.insert_to_treeview)
        self.insert_button.grid(row=4, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")

        self.delete_button = ttk.Button(left_frame, text="Xóa", command=self.delete_selected_row)
        self.delete_button.grid(row=5, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")

        self.export_button = ttk.Button(left_frame, text="Xuất Excel", command=self.export_to_excel)
        self.export_button.grid(row=6, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")

        # Đảm bảo các cột trong frame có tỷ lệ hợp lý
        left_frame.columnconfigure(1, weight=1)

        # Tạo frame chứa Treeview
        tree_frame = ttk.Frame(self.root)  # Sửa: Thay 'frame' bằng 'self.root' hoặc một frame hợp lệ
        tree_frame.place(x=250, y=410, width=950, height=250)  # Đặt vị trí và kích thước cho frame

        # Thêm thanh cuộn dọc
        tree_scroll = ttk.Scrollbar(tree_frame, orient="vertical")
        tree_scroll.pack(side="right", fill="y")

        # Định nghĩa các cột cho Treeview
        cols = ("STT", "Name", "Author", "Year", "Rating", "Genre")
        self.treeview = ttk.Treeview(tree_frame, show="headings", columns=cols, yscrollcommand=tree_scroll.set)

        # Cấu hình tiêu đề và kích thước cột
        self.treeview.heading("STT", text="STT")
        self.treeview.column("STT", width=50, anchor="center")  # Cột số thứ tự nhỏ hơn và căn giữa

        for col in cols[1:]:  # Bỏ qua cột "STT" vì đã cấu hình riêng
            self.treeview.heading(col, text=col)
            self.treeview.column(col, width=150)
        
        # Hiển thị Treeview
        self.treeview.pack(fill=BOTH, expand=True)
        
        # Kết nối thanh cuộn với Treeview
        tree_scroll.config(command=self.treeview.yview)

    def insert_data(self):
        """Hàm xử lý khi nhấn nút 'Thêm'."""
        print("Thêm dữ liệu...")

    def delete_data(self):
        """Hàm xử lý khi nhấn nút 'Xóa'."""
        print("Xóa dữ liệu...")

    def export_data(self):
        """Hàm xử lý khi nhấn nút 'Xuất Excel'."""
        print("Xuất dữ liệu ra Excel...")

    def center_window(self, width, height):
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x_position = (screen_width - width) // 2
        y_position = (screen_height - height) // 2 - 40  # Giảm 20px để căn giữa chính xác hơn
        self.root.geometry(f"{width}x{height}+{x_position}+{y_position}")


    def create_ui(self):
        icon_image = PhotoImage(file="images/icon.png")
        self.root.iconphoto(False, icon_image)

        # Tạo Canvas
        canvas = Canvas(self.root, width=1250, height=225, bg="#FFFFCC", highlightthickness=0)
        canvas.place(x=0, y=0)
        create_rounded_rectangle(canvas, 10, 10, 1240, 160, radius=20, fill="#FF6666", outline="")

        # Logo
        logo_image = PhotoImage(file="images/logo.png")
        Label(self.root, image=logo_image, bg="#FF6666").place(x=100, y=80)
        self.root.logo_image = logo_image

        # Tiêu đề chính
        heading = Label(self.root, text="BOOK RECOMMENDATION SYSTEM", font=("Lato", 30, "bold"), fg='white', bg="#FF6666")
        heading.place(x=280, y=20)

        # Thanh tìm kiếm
        search_box_image = PhotoImage(file="images/Rectangle_2.png")
        Label(self.root, image=search_box_image, bg="#FF6666").place(x=294, y=80)
        self.root.search_box_image = search_box_image

        self.search_var = StringVar()
        search_entry = Entry(self.root, textvariable=self.search_var, width=20, font=("Lato", 25), bg='white', fg='black', bd=0)
        search_entry.place(x=415, y=98)

        recommend_button_image = PhotoImage(file="images/Search.png")
        recommend_button = Button(self.root, image=recommend_button_image, bg="#FF6666", bd=0, cursor='hand2', command=self.search_books)
        recommend_button.place(x=860, y=94)
        self.root.recommend_button_image = recommend_button_image

        # Bắt sự kiện phím Enter
        search_entry.bind("<Return>", lambda event: self.search_books())

        recommend_button_image = PhotoImage(file="images/Search.png")
        recommend_button = Button(self.root, image=recommend_button_image, bg="#FF6666", bd=0, cursor='hand2', command=self.search_books)
        recommend_button.place(x=860, y=94)
        self.root.recommend_button_image = recommend_button_image

        # Nút cài đặt
        setting_image = PhotoImage(file="images/setting.png")
        setting_button = Button(self.root, image=setting_image, bd=0, cursor='hand2', bg="#FF6666")
        setting_button.place(x=1050, y=95)
        self.root.setting_image = setting_image

        setting_menu = Menu(self.root, tearoff=0)
        setting_menu.add_checkbutton(label="Publish Date", variable=self.check_var_date)
        setting_menu.add_checkbutton(label="Rating", variable=self.check_var_rating)
        setting_button.bind('<Button-1>', lambda event: setting_menu.post(event.x_root, event.y_root))

        # Nút đăng xuất
        logout_image = PhotoImage(file="images/logout.png")
        Button(self.root, image=logout_image, bg="#FF6666", cursor='hand2', command=self.root.destroy).place(x=1150, y=20)
        self.root.logout_image = logout_image

        # Nút quản lý dữ liệu
        Button(self.root, text="Data", font=("Lato", 15, "bold"), bg="#FF6666", fg="white", command=self.open_data_management).place(x=300, y=230)

        # Gọi hàm để tạo giao diện hiển thị sách với bo góc
        self.frames, self.text_labels, self.image_labels = create_book_display(self.root)


    def open_data_management(self):
        BookDataManagement(self.root)

    
    def fetch_information(self, title, poster, date, rating):
        """
        Cập nhật thông tin sách vào giao diện hiển thị.
        """
        if self.inc >= len(self.frames):
            return

        text_label = self.text_labels[self.inc]
        image_label = self.image_labels[self.inc]

        text_label.config(text=title)

        if self.check_var_date.get():
            text_label.config(text=f"{title}\nDate: {date}")

        if self.check_var_rating.get():
            text_label.config(text=f"{title}\nRating: {rating}")

        if poster != 'N/A':
            try:
                response = requests.get(poster, timeout=5)
                response.raise_for_status()
                img_data = response.content
                img = Image.open(BytesIO(img_data))
                resized_image = img.resize((120, 160))
                photo = ImageTk.PhotoImage(resized_image)
                image_label.config(image=photo)
                image_label.image = photo  # Sửa: Lưu tham chiếu để tránh bị xóa
            except requests.exceptions.RequestException as e:
                print(f"Lỗi khi tải ảnh: {e}")

        self.inc += 1

    # def search_books(self):
    #     """Tìm kiếm sách từ Google Books API và hiển thị kết quả."""
    #     start_time = time.time()

    #     self.inc = 0
    #     search_query = self.search_var.get().strip()

    #     if not search_query:
    #         messagebox.showinfo("Thông báo", "Vui lòng nhập từ khóa tìm kiếm!")
    #         return

    #     search_query_encoded = urllib.parse.quote(search_query)
    #     url = f"https://www.googleapis.com/books/v1/volumes?q={search_query_encoded}&maxResults=8"

    #     try:
    #         response = requests.get(url, timeout=10)
    #         response.raise_for_status()
    #         data = response.json()

    #         if "items" not in data:
    #             messagebox.showinfo("Không tìm thấy", "Không có kết quả phù hợp.")
    #             return

    #         for item in data["items"]:
    #             volume_info = item.get("volumeInfo", {})
    #             title = volume_info.get("title", "N/A")
    #             published_date = volume_info.get("publishedDate", "N/A")
    #             rating = volume_info.get("averageRating", "N/A")
    #             poster = volume_info.get("imageLinks", {}).get("thumbnail", "N/A")

    #             self.fetch_information(title, poster, published_date, rating)

    #     except requests.exceptions.RequestException as e:
    #         messagebox.showerror("Lỗi", f"Lỗi khi lấy dữ liệu từ API: {e}")

    #     end_time = time.time()
    #     print(f"⏳ Thời gian gọi API: {end_time - start_time:.2f} giây")
    #     print(f"📚 Số sách hiển thị: {self.inc}")
    def search_books(self):
        """Tìm kiếm sách từ Google Books API và hiển thị kết quả."""
        start_time = time.time()

        # Hiển thị biểu tượng "Loading..."
        print("⏳ Đang tải dữ liệu...")
        self.root.update()  # Cập nhật giao diện để hiển thị biểu tượng "Loading..."
        loading_label = Label(self.root, text="Loading...", font=("Arial", 15, "bold"), fg="black", bg="yellow")
        loading_label.place(x=50, y=20)  # Đặt vị trí biểu tượng "Loading"

        self.inc = 0
        search_query = self.search_var.get().strip()

        if not search_query:
            loading_label.destroy()  # Xóa biểu tượng "Loading" nếu không có từ khóa
            messagebox.showinfo("Thông báo", "Vui lòng nhập từ khóa tìm kiếm!")
            return

        search_query_encoded = urllib.parse.quote(search_query)
        url = f"https://www.googleapis.com/books/v1/volumes?q={search_query_encoded}&maxResults=8"

        try:
            response = requests.get(url, timeout=10)
            response.raise_for_status()
            data = response.json()

            if "items" not in data:
                loading_label.destroy()  # Xóa biểu tượng "Loading" nếu không có kết quả
                messagebox.showinfo("Không tìm thấy", "Không có kết quả phù hợp.")
                return

            for item in data["items"]:
                volume_info = item.get("volumeInfo", {})
                title = volume_info.get("title", "N/A")
                published_date = volume_info.get("publishedDate", "N/A")
                rating = volume_info.get("averageRating", "N/A")
                poster = volume_info.get("imageLinks", {}).get("thumbnail", "N/A")

                self.fetch_information(title, poster, published_date, rating)

        except requests.exceptions.RequestException as e:
            loading_label.destroy()  # Xóa biểu tượng "Loading" nếu có lỗi
            messagebox.showerror("Lỗi", f"Lỗi khi lấy dữ liệu từ API: {e}")

        # Xóa biểu tượng "Loading" sau khi hoàn tất
        loading_label.destroy()

        end_time = time.time()
        print(f"⏳ Thời gian gọi API: {end_time - start_time:.2f} giây")
        print(f"📚 Số sách hiển thị: {self.inc}")

    def insert_to_treeview(self):
        """Thêm dữ liệu vào Treeview."""
        name = self.name_entry.get().strip()
        author = self.author_spinbox.get().strip()
        rating = self.rating_combobox.get().strip()
        employment_status = "Yes" if self.employed_var.get() else "No"

        if not name or not author:
            messagebox.showwarning("Cảnh báo", "Vui lòng nhập tên sách và tác giả hợp lệ!")
            return

        # Tính số thứ tự dựa trên số lượng hàng hiện tại
        stt = len(self.treeview.get_children()) + 1

        # Thêm dữ liệu vào Treeview
        self.treeview.insert("", "end", values=(stt, name, author, rating, employment_status))
        self.name_entry.delete(0, "end")
        self.name_entry.insert(0, "Tên sách")
        self.author_spinbox.delete(0, "end")
        self.author_spinbox.insert(0, "1")
        self.rating_combobox.current(0)
        self.employed_var.set(False)

    def delete_selected_row(self):
        """Xóa hàng được chọn trong Treeview và cập nhật lại số thứ tự."""
        selected_item = self.treeview.selection()
        if not selected_item:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn hàng cần xóa!")
            return

        # Xóa hàng được chọn
        self.treeview.delete(selected_item)

        # Cập nhật lại số thứ tự
        for index, child in enumerate(self.treeview.get_children(), start=1):
            values = self.treeview.item(child)["values"]
            self.treeview.item(child, values=(index, *values[1:]))

    def export_to_excel(self):
        """Xuất dữ liệu từ Treeview ra tệp Excel (.xlsx)."""
        # Chọn nơi lưu file
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not path:
            return  # Nếu không chọn file, thoát hàm

        # Kiểm tra dữ liệu trong Treeview
        if not self.treeview.get_children():
            messagebox.showwarning("Cảnh báo", "Không có dữ liệu để xuất!")
            return

        try:
            # Tạo workbook và sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Data"

            # Thêm tiêu đề cột
            cols = ("STT", "Name", "Author", "Rating", "Genre")
            for col_num, col_name in enumerate(cols, start=1):
                cell = sheet.cell(row=1, column=col_num, value=col_name)
                cell.font = openpyxl.styles.Font(bold=True)  # In đậm tiêu đề
                cell.alignment = openpyxl.styles.Alignment(horizontal="center")  # Căn giữa

            # Thêm dữ liệu từ Treeview
            for row_num, child in enumerate(self.treeview.get_children(), start=2):
                values = self.treeview.item(child)["values"]
                for col_num, value in enumerate(values, start=1):
                    sheet.cell(row=row_num, column=col_num, value=value)

            # Lưu file
            workbook.save(path)
            messagebox.showinfo("Thông báo", f"Dữ liệu đã được xuất ra tệp Excel tại {path}")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xuất dữ liệu: {e}")